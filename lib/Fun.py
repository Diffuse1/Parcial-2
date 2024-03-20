import openpyxl
from.Classes import nodo
#--------------------Arbol de Extensión-------------------------------------
def Datos (Archivo):
    worbook = openpyxl.load_workbook(Archivo)
    Hoja = worbook.active
    data = [ ]
    for row in Hoja.iter_rows(values_only=True):
        data.append(row)
    return data

def Red(data,numero):
    for row in data:
        for celda in row:
            if celda == numero:
                print("\033[34m{celda}\033[0m".format(celda), end='')
            else:
                print(celda, end=' ')
        print()
#-------------------Binario-------------------------------------------------------------------------
def cargar(Archivo):
    workbook = openpyxl.load_workbook(Archivo)
    sheet = workbook.active
    datos = [cell for row in sheet.iter_rows(values_only=True) for cell in row]
    return datos

def arbol_ordenado(datos):
    raiz = nodo(datos[0])
    for dato in datos[1:]:
        nuevo_nodo = nodo(dato)
        nodos_ordenados(raiz, nuevo_nodo)
    return raiz

def nodos_ordenados(nodo_padre, nuevo_nodo):
    if nuevo_nodo.valor < nodo_padre.valor:
        if nodo_padre.izq is None:
            nodo_padre.izq = nuevo_nodo
        else:
            nodos_ordenados(nodo_padre.izq, nuevo_nodo)
    elif nuevo_nodo.valor > nodo_padre.valor:
        if nodo_padre.der is None:
            nodo_padre.der = nuevo_nodo
        else:
            nodos_ordenados(nodo_padre.der, nuevo_nodo)

def print_arbol(nodo):
    if nodo is not None:
        print(nodo.get_arbol())
        print_arbol(nodo.izq)        
        print_arbol(nodo.der)

def LVR(Nodo,):
    inOrderArr = []
    if Nodo:
        inOrderArr += LVR(Nodo.izq)
        inOrderArr.append(Nodo.valor)
        inOrderArr += LVR(Nodo.der)
    return inOrderArr

def VLR(Nodo):
    preOrderArr = []
    if Nodo:
        preOrderArr.append(Nodo.valor)
        preOrderArr += VLR(Nodo.izq)
        preOrderArr += VLR(Nodo.der)
    return preOrderArr

def LRV(Nodo):
    postOrderArr = []
    if Nodo:
        postOrderArr += LRV(Nodo.izq)
        postOrderArr += LRV(Nodo.der)
        postOrderArr.append(Nodo.valor)
    return postOrderArr
#_---------------------------No binario-------------------------------------
def Ordenado(Nodo,):
    OrdenadoArr = []
    if Nodo:
        OrdenadoArr += Ordenado(Nodo.izq)
        OrdenadoArr.append(Nodo.valor)
        OrdenadoArr += Ordenado(Nodo.der)
    return OrdenadoArr


Archivo = input('Nombre del archivo "Eval.xlsm:" ')
datos = cargar(Archivo)
arbol = arbol_ordenado(datos)
numero = input('Ingresa tu Número: ')
data = Datos(Archivo)
Rojo = Red(data,numero)
    